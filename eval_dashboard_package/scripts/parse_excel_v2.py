#!/usr/bin/env python3
"""Parse Excel and build case data with reply texts for Case Explorer"""
import openpyxl, json

wb = openpyxl.load_workbook('cases_excel.xlsx', read_only=True, data_only=True)
ws = wb['评估数据']

# Column mapping (1-indexed):
# 1: 评估轮次, 2: session题号, 3: prompt题号, 4: prompt
# 5: 评估方向, 6: 单轮/多轮, 7: source
# 11: 豆包回复文本, 27: Qwen回复文本
# 12: 豆包客观dcg(prompt), 18: 豆包主客观dcg(prompt)
# 30: Qwen客观dcg(prompt), 36: Qwen主客观dcg(prompt)
# 45: GSB客观(prompt), 51: GSB主客观(prompt)
# 57: 豆包字数, 58: Qwen字数

cases = []
for r in range(4, ws.max_row + 1):
    prompt_id = ws.cell(r, 3).value
    if not prompt_id:
        continue
    
    prompt_text = str(ws.cell(r, 4).value or '')
    # Clean up JSON-like prompts (multimedia)
    if prompt_text.startswith('{"entities"'):
        prompt_text = '[多媒体内容]'
    
    cat = str(ws.cell(r, 5).value or '').strip()
    if not cat or cat == 'None':
        cat = '其他'
    cat = cat.replace('VLM（通用）', 'VLM通用').replace('VLM（教育）', 'VLM教育')
    
    def safe_str(v):
        if v is None:
            return ''
        return str(v).strip()
    
    def safe_int(v):
        if v is None:
            return 0
        try:
            return int(v)
        except:
            return 0
    
    # Get reply texts (truncate to 500 chars for display)
    db_reply = safe_str(ws.cell(r, 11).value)
    qw_reply = safe_str(ws.cell(r, 27).value)
    
    c = {
        'id': safe_str(prompt_id),
        'sid': safe_str(ws.cell(r, 2).value),
        'r': safe_int(ws.cell(r, 1).value),
        'q': prompt_text,
        'cat': cat,
        'src': safe_str(ws.cell(r, 7).value) or 'APP',
        'turn': safe_str(ws.cell(r, 6).value),
        'db_obj': safe_str(ws.cell(r, 12).value),
        'db_sub': safe_str(ws.cell(r, 18).value),
        'qw_obj': safe_str(ws.cell(r, 30).value),
        'qw_sub': safe_str(ws.cell(r, 36).value),
        'gsb_obj': safe_str(ws.cell(r, 45).value),
        'gsb_sub': safe_str(ws.cell(r, 51).value),
        'db_w': safe_int(ws.cell(r, 57).value),
        'qw_w': safe_int(ws.cell(r, 58).value),
        'db_r': db_reply[:800] if len(db_reply) > 800 else db_reply,
        'qw_r': qw_reply[:800] if len(qw_reply) > 800 else qw_reply,
    }
    
    # Compute GSB label
    g = c['gsb_obj']
    try:
        v = int(g)
        if v > 0:
            c['gsb_label'] = '豆包胜 (+' + str(v) + ')'
            c['gsb_cls'] = 'gsb-win'
        elif v < 0:
            c['gsb_label'] = 'Qwen胜 (' + str(v) + ')'
            c['gsb_cls'] = 'gsb-lose'
        else:
            c['gsb_label'] = '平局'
            c['gsb_cls'] = 'gsb-tie'
    except:
        c['gsb_label'] = ''
        c['gsb_cls'] = ''
    
    cases.append(c)

print(f"Parsed {len(cases)} cases")

# Stats
cats = {}
for c in cases:
    cats[c['cat']] = cats.get(c['cat'], 0) + 1
print("Categories:", sorted(cats.items(), key=lambda x: -x[1]))

srcs = {}
for c in cases:
    srcs[c['src']] = srcs.get(c['src'], 0) + 1
print("Sources:", srcs)

# Check reply text coverage
db_has_reply = sum(1 for c in cases if c['db_r'])
qw_has_reply = sum(1 for c in cases if c['qw_r'])
print(f"豆包回复文本: {db_has_reply}/{len(cases)}")
print(f"Qwen回复文本: {qw_has_reply}/{len(cases)}")

with open('cases_data_v2.json', 'w', encoding='utf-8') as f:
    json.dump(cases, f, ensure_ascii=False, separators=(',', ':'))

print(f"Saved cases_data_v2.json: {len(open('cases_data_v2.json','rb').read())} bytes")
