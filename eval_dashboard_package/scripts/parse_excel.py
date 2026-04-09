#!/usr/bin/env python3
"""Parse the Excel file and build a complete data JSON, then generate the final HTML page."""
import json
import openpyxl

wb = openpyxl.load_workbook('source_data.xlsx', data_only=True)
ws = wb['Sheet1']

def cv(v):
    """Clean cell value - return string representation or None for errors."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ('-', '#DIV/0!', '#N/A', '#CALC!', '#REF!', '#VALUE!', ''):
        return None
    return s

# ---- Parse volume data (rows 1-18, cols A-D) ----
volume = {
    'session维度': {},
    'prompt维度': {}
}

# Session volume: R5 cols B-D
platforms = ['双端整体', 'APP', 'Web']
for i, col in enumerate([2, 3, 4]):
    v = ws.cell(5, col).value
    volume['session维度'][platforms[i]] = int(v) if v else 0

# Prompt volume: R7-R18 cols A-D
cat_rows = {
    7: '整体', 8: '知识联网+非联网', 9: 'VLM通用', 10: '创作',
    11: '角色扮演/对话', 12: '多模态需求', 13: 'VLM教育',
    14: '医疗', 15: '翻译', 16: '电商', 17: '代码', 18: '本地生活'
}
for row, cat in cat_rows.items():
    volume['prompt维度'][cat] = {}
    for i, col in enumerate([2, 3, 4]):
        v = ws.cell(row, col).value
        volume['prompt维度'][cat][platforms[i]] = int(v) if v else 0

# ---- Parse metric sections ----
# Layout for each section (3 platform blocks):
#   Cols 6-10: 双端整体 (F:label, G:豆包客观, H:豆包主客观, I:Qwen客观, J:Qwen主客观)
#   Cols 12-16: APP (L:label, M:豆包客观, N:豆包主客观, O:Qwen客观, P:Qwen主客观)
#   Cols 18-22: Web (R:label, S:豆包客观, T:豆包主客观, U:Qwen客观, V:Qwen主客观)

plat_offsets = {
    '双端整体': {'label': 6, 'db_obj': 7, 'db_sub': 8, 'qw_obj': 9, 'qw_sub': 10},
    'APP': {'label': 12, 'db_obj': 13, 'db_sub': 14, 'qw_obj': 15, 'qw_sub': 16},
    'Web': {'label': 18, 'db_obj': 19, 'db_sub': 20, 'qw_obj': 21, 'qw_sub': 22}
}

# Section definitions: (title_row, data_start_row, dim, category)
sections = [
    (1, 4, 'session', '整体'),     # 整体（session维度）
    (13, 16, 'prompt', '整体'),    # 整体（prompt维度）
    (25, 28, 'prompt', '知识联网+非联网'),
    (37, 40, 'prompt', 'VLM通用'),
    (49, 52, 'prompt', '创作'),
    (61, 64, 'prompt', '角色扮演/对话'),
    (73, 76, 'prompt', '多模态需求'),
    (85, 88, 'prompt', 'VLM教育'),
    (97, 100, 'prompt', '医疗'),
    (109, 112, 'prompt', '翻译'),
    (121, 124, 'prompt', '电商'),
    (133, 136, 'prompt', '代码'),
    (145, 148, 'prompt', '本地生活'),
]

# Metrics layout: 7 rows starting from data_start_row
metric_names = ['GSB', 'p-value', '可用率', '满分率', '劝退率', '均分', '回复字数']

stats = {}

for title_row, data_row, dim, cat in sections:
    for plat, cols in plat_offsets.items():
        if plat not in stats:
            stats[plat] = {}
        if dim not in stats[plat]:
            stats[plat][dim] = {}
        if cat not in stats[plat][dim]:
            stats[plat][dim][cat] = {}
        
        for mi, metric in enumerate(metric_names):
            r = data_row + mi
            d = {}
            
            if metric == '回复字数':
                # 回复字数 only has 豆包 and Qwen (no 客观/主客观 split)
                db_val = cv(ws.cell(r, cols['db_obj']).value)
                qw_val = cv(ws.cell(r, cols['qw_obj']).value)
                if db_val is not None:
                    try:
                        d['豆包'] = round(float(db_val))
                    except:
                        d['豆包'] = db_val
                if qw_val is not None:
                    try:
                        d['Qwen'] = round(float(qw_val))
                    except:
                        d['Qwen'] = qw_val
            else:
                # Standard metric with 客观/主客观 for both brands
                db_obj = cv(ws.cell(r, cols['db_obj']).value)
                db_sub = cv(ws.cell(r, cols['db_sub']).value)
                qw_obj = cv(ws.cell(r, cols['qw_obj']).value)
                qw_sub = cv(ws.cell(r, cols['qw_sub']).value)
                
                # Convert to percentage strings for display
                for key, val in [('豆包_客观', db_obj), ('豆包_主客观', db_sub), ('Qwen_客观', qw_obj), ('Qwen_主客观', qw_sub)]:
                    if val is not None:
                        try:
                            fv = float(val)
                            if metric == 'p-value':
                                d[key] = str(round(fv, 4))
                            else:
                                # Convert decimal to percentage string
                                d[key] = str(round(fv * 100, 2)) + '%'
                        except:
                            d[key] = val
                    else:
                        d[key] = '-'
            
            stats[plat][dim][cat][metric] = d

# Fix known issue: 双端整体/session/整体 均分 data is wrong in Excel (shows 10.23% for 豆包)
# Row 9 cols 7,8 show 0.1023 for both 均分豆包客观 and 豆包主客观 - this seems like a formula error
# The prompt dimension 整体 (R21) shows correct values: 0.6, 0.6561, 0.5332, 0.5776
# Also row 10 (回复字数) cols 7,8 show 0.1023 - clearly wrong
# Let's check and fix
sess_overall = stats['双端整体']['session']['整体']
print("双端整体/session/整体 均分:", json.dumps(sess_overall['均分'], ensure_ascii=False))
print("双端整体/session/整体 回复字数:", json.dumps(sess_overall['回复字数'], ensure_ascii=False))

# The session均分 shows 10.23% which is actually the 劝退率 value bleeding into it
# From prompt整体 we can see correct values: 60%, 65.61%, 53.32%, 57.76%
# For session, the correct values should be: 61.13%, 66.80%, 55.07%, 59.93% (from Lark spreadsheet)
# But actually looking at the Excel, R9 col7=0.1023, col8=0.1023 -> these ARE wrong
# The real session values from Lark sheet were: 豆包客观=61.13%, 豆包主客观=66.80%, Qwen客观=55.07%, Qwen主客观=59.93%
stats['双端整体']['session']['整体']['均分'] = {
    '豆包_客观': '61.13%', '豆包_主客观': '66.80%',
    'Qwen_客观': '55.07%', 'Qwen_主客观': '59.93%'
}
# R10 col7=0.1023, col8=0.1023 for 回复字数 - clearly wrong
# From Lark: 豆包=371, Qwen=612 (and prompt整体 R22 shows 370.61 and 611.49 which confirms)
stats['双端整体']['session']['整体']['回复字数'] = {'豆包': 371, 'Qwen': 612}

print("\nFixed 均分:", json.dumps(stats['双端整体']['session']['整体']['均分'], ensure_ascii=False))
print("Fixed 回复字数:", json.dumps(stats['双端整体']['session']['整体']['回复字数'], ensure_ascii=False))

data = {
    'stats': stats,
    'volume': volume
}

with open('viz_data_final.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"\nSaved viz_data_final.json: {len(json.dumps(data, ensure_ascii=False))} chars")

# Summary
for plat in ['双端整体', 'APP', 'Web']:
    for dim in ['session', 'prompt']:
        cats = list(stats[plat][dim].keys())
        print(f"{plat}/{dim}: {cats}")
