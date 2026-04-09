#!/usr/bin/env python3
"""
一键构建竞对可视化页面
用法: python3 auto_build.py <excel_file>
输出: viz_output.html

支持两种 Excel 格式：
- 旧版(58列): 无 message_id 列，回复文本在 col 10
- 新版(62列): 有 message_id 列，回复文本在 col 12
自动检测并适配。同时从"数据统计"sheet解析dashboard DATA。
"""

import openpyxl, json, re, os, sys

# ============================================================
# CONFIG: Column mapping (auto-detected)
# ============================================================
class ColMap:
    """Column index mapping for both old and new Excel formats."""
    def __init__(self, ws):
        h0 = ws.cell(3, 1).value
        if h0 == 'message_id':
            self.fmt = 'new'
            self._init_new()
        else:
            self.fmt = 'old'
            self._init_old()
        print('Detected format: %s (%d-col)' % (self.fmt, 62 if self.fmt == 'new' else 58))

    def _init_old(self):
        self.sid = 1; self.pid = 2; self.prompt = 3
        self.cat = 4; self.turn = 5; self.src = 6
        self.recall = 7; self.rich_intent = 8
        self.db_reply = 10
        self.db_obj_p = 11; self.db_obj_p_qt = 12; self.db_obj_p_note = 13
        self.db_obj_s = 14; self.db_obj_s_qt = 15; self.db_obj_s_note = 16
        self.db_sub_p = 17; self.db_sub_p_qt = 18; self.db_sub_p_note = 19
        self.db_sub_s = 20; self.db_sub_s_qt = 21; self.db_sub_s_note = 22
        self.db_rich_s = 23; self.db_rich_t = 24; self.db_rich_n = 25
        self.qw_reply = 26; self.qw_cot = 27; self.qw_attach = 28
        self.qw_obj_p = 29; self.qw_obj_p_qt = 30; self.qw_obj_p_note = 31
        self.qw_obj_s = 32; self.qw_obj_s_qt = 33; self.qw_obj_s_note = 34
        self.qw_sub_p = 35; self.qw_sub_p_qt = 36; self.qw_sub_p_note = 37
        self.qw_sub_s = 38; self.qw_sub_s_qt = 39; self.qw_sub_s_note = 40
        self.qw_rich_s = 41; self.qw_rich_t = 42; self.qw_rich_n = 43
        self.gsb_obj_p = 44; self.gsb_obj_p_dt = 45; self.gsb_obj_p_note = 46
        self.gsb_obj_s = 47; self.gsb_obj_s_dt = 48; self.gsb_obj_s_note = 49
        self.gsb_sub_p = 50; self.gsb_sub_p_dt = 51; self.gsb_sub_p_note = 52
        self.gsb_sub_s = 53; self.gsb_sub_s_dt = 54; self.gsb_sub_s_note = 55
        self.db_wc = 56; self.qw_wc = 57

    def _init_new(self):
        self.sid = 2; self.pid = 3; self.prompt = 4
        self.cat = 5; self.turn = 6; self.src = 7
        self.recall = 8; self.rich_intent = 9
        self.db_reply = 12
        self.db_obj_p = 13; self.db_obj_p_qt = 14; self.db_obj_p_note = 15
        self.db_obj_s = 16; self.db_obj_s_qt = 17; self.db_obj_s_note = 18
        self.db_sub_p = 19; self.db_sub_p_qt = 20; self.db_sub_p_note = 21
        self.db_sub_s = 22; self.db_sub_s_qt = 23; self.db_sub_s_note = 24
        self.db_rich_s = 25; self.db_rich_t = 26; self.db_rich_n = 27
        self.qw_reply = 30; self.qw_cot = 31; self.qw_attach = 32
        self.qw_obj_p = 33; self.qw_obj_p_qt = 34; self.qw_obj_p_note = 35
        self.qw_obj_s = 36; self.qw_obj_s_qt = 37; self.qw_obj_s_note = 38
        self.qw_sub_p = 39; self.qw_sub_p_qt = 40; self.qw_sub_p_note = 41
        self.qw_sub_s = 42; self.qw_sub_s_qt = 43; self.qw_sub_s_note = 44
        self.qw_rich_s = 45; self.qw_rich_t = 46; self.qw_rich_n = 47
        self.gsb_obj_p = 48; self.gsb_obj_p_dt = 49; self.gsb_obj_p_note = 50
        self.gsb_obj_s = 51; self.gsb_obj_s_dt = 52; self.gsb_obj_s_note = 53
        self.gsb_sub_p = 54; self.gsb_sub_p_dt = 55; self.gsb_sub_p_note = 56
        self.gsb_sub_s = 57; self.gsb_sub_s_dt = 58; self.gsb_sub_s_note = 59
        self.db_wc = 60; self.qw_wc = 61


# ============================================================
# HELPERS
# ============================================================
def safe_str(v):
    if v is None: return ''
    s = str(v)
    if s == 'NULL' or s == 'null': return ''
    return s

def parse_prompt(raw):
    if not raw: return '', []
    s = str(raw)
    imgs = []
    try:
        obj = json.loads(s)
        if isinstance(obj, dict) and 'entities' in obj:
            text = obj.get('text', '')
            for ent in obj.get('entities', []):
                if ent.get('entity_type') == 2:
                    ec = ent.get('entity_content', {})
                    img = ec.get('image', {})
                    ori = img.get('image_ori', {})
                    url = ori.get('url', '')
                    if url: imgs.append(url)
            return text, imgs
    except:
        pass
    return s, []

def sanitize_html(s):
    if not s: return ''
    s = str(s)
    s = s.replace('<script', '<scr ipt').replace('</script', '</scr ipt')
    s = s.replace('<Script', '<Scr ipt').replace('</Script', '</Scr ipt')
    return s

def fmt_pct(v):
    """Format a float as '12.34%', or return '-' for non-numeric."""
    if v is None: return '-'
    s = str(v)
    if s == '-' or s.startswith('#') or not s: return '-'
    try:
        f = float(v)
        return str(round(f * 100, 2)) + '%'
    except:
        return '-'

def fmt_int(v):
    """Format as integer or return 0."""
    if v is None: return 0
    try:
        return int(round(float(v)))
    except:
        return 0

def fmt_pval(v):
    """Format p-value."""
    if v is None: return '-'
    s = str(v)
    if s == '-' or s.startswith('#'): return '-'
    try:
        f = float(v)
        if f < 0.0001: return '0.0'
        return str(round(f, 4))
    except:
        return '-'

def fmt_wc(v):
    """Format word count as int."""
    if v is None: return 0
    s = str(v)
    if s == '-' or s.startswith('#'): return 0
    try:
        return int(round(float(v)))
    except:
        return 0


# ============================================================
# STEP 1: Parse Excel -> case JSON
# ============================================================
def parse_cases(excel_path, wb):
    ws = wb['评估数据']
    cm = ColMap(ws)

    cases = []
    for ri, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        vals = list(row)
        prompt_raw = vals[cm.prompt] if cm.prompt < len(vals) else None
        if not prompt_raw: continue

        prompt_text, imgs = parse_prompt(prompt_raw)
        if not prompt_text and not imgs: continue

        def g(idx):
            return safe_str(vals[idx]) if idx < len(vals) else ''

        pid = g(cm.pid)
        case_id = pid if pid else 'R' + str(ri)

        # GSB: positive = 豆包胜
        gsb_obj = g(cm.gsb_obj_p)
        gsb_sub = g(cm.gsb_sub_p)
        gsb_label = ''
        gsb_cls = 'gsb-tie'
        try:
            gv = float(gsb_obj) if gsb_obj else None
            if gv is not None:
                gvi = int(gv) if gv == int(gv) else gv
                if gv > 0:
                    gsb_label = '\u8c46\u5305\u80dc (+%s)' % str(gvi)
                    gsb_cls = 'gsb-win'
                elif gv < 0:
                    gsb_label = 'Qwen\u80dc (%s)' % str(gvi)
                    gsb_cls = 'gsb-lose'
                else:
                    gsb_label = '\u5e73\u5c40 (0)'
        except:
            pass

        rec = {
            'id': case_id, 'sid': g(cm.sid), 'r': ri - 3,
            'q': prompt_text, 'cat': g(cm.cat), 'src': g(cm.src), 'turn': g(cm.turn),
            'db_obj': g(cm.db_obj_p), 'db_sub': g(cm.db_sub_p),
            'qw_obj': g(cm.qw_obj_p), 'qw_sub': g(cm.qw_sub_p),
            'gsb_obj': gsb_obj, 'gsb_sub': gsb_sub,
            'gsb_label': gsb_label, 'gsb_cls': gsb_cls,
            'db_w': g(cm.db_wc), 'qw_w': g(cm.qw_wc),
            'db_r': sanitize_html(vals[cm.db_reply] if cm.db_reply < len(vals) else None),
            'qw_r': sanitize_html(vals[cm.qw_reply] if cm.qw_reply < len(vals) else None),
            'imgs': imgs,
            'db_oq': g(cm.db_obj_p_qt), 'db_on': g(cm.db_obj_p_note),
            'db_sq': g(cm.db_sub_p_qt), 'db_sn': g(cm.db_sub_p_note),
            'db_rs': g(cm.db_rich_s), 'db_rt': g(cm.db_rich_t), 'db_rn': g(cm.db_rich_n),
            'qw_oq': g(cm.qw_obj_p_qt), 'qw_on': g(cm.qw_obj_p_note),
            'qw_sq': g(cm.qw_sub_p_qt), 'qw_sn': g(cm.qw_sub_p_note),
            'qw_rs': g(cm.qw_rich_s), 'qw_rt': g(cm.qw_rich_t), 'qw_rn': g(cm.qw_rich_n),
            'go_dt': g(cm.gsb_obj_p_dt), 'go_n': g(cm.gsb_obj_p_note),
            'gs_dt': g(cm.gsb_sub_p_dt), 'gs_n': g(cm.gsb_sub_p_note),
        }
        cases.append(rec)

    print('Parsed %d cases' % len(cases))
    return cases


# ============================================================
# STEP 2: Parse 数据统计 sheet -> DATA for dashboard
# ============================================================
def parse_dashboard_data(wb):
    """
    Parse the 数据统计 sheet into the DATA dict matching the HTML dashboard format.
    DATA = {
      "stats": {
        "双端整体": {
          "session": { "<cat>": {<metric>: {<key>: val}} },
          "prompt": { "<cat>": {<metric>: {<key>: val}} }
        },
        "APP": { ... },
        "Web": { ... }
      },
      "volume": {
        "session维度": { "双端整体": N, "APP": N, "Web": N },
        "prompt维度": { "<cat>": {"双端整体": N, "APP": N, "Web": N} }
      }
    }
    """
    ws = wb['数据统计']
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(list(row))

    nrows = len(all_rows)

    # ---- Volume section (rows 0-17 area, col 0-3) ----
    # Row 2 (0-indexed): "session维度"
    # Row 3: "整体" with session counts
    # Row 4: "prompt维度"
    # Rows 5-16: category prompt counts
    volume = {}
    # Session counts
    # Row index 3 (0-indexed) = "整体" session counts
    session_vol = {}
    r3 = all_rows[3]  # 整体 session
    session_vol['双端整体'] = fmt_int(r3[1])
    session_vol['APP'] = fmt_int(r3[2])
    session_vol['Web'] = fmt_int(r3[3])
    volume['session维度'] = session_vol

    # Prompt counts (rows 5-16, 0-indexed)
    prompt_vol = {}
    categories_in_order = []
    for ri in range(5, 18):
        if ri >= nrows: break
        r = all_rows[ri]
        cat_name = r[0]
        if not cat_name: continue
        cat_name = str(cat_name)
        categories_in_order.append(cat_name)
        prompt_vol[cat_name] = {
            '双端整体': fmt_int(r[1]),
            'APP': fmt_int(r[2]),
            'Web': fmt_int(r[3])
        }
    volume['prompt维度'] = prompt_vol

    # ---- Stats section ----
    # Layout in cols 5-9 = 双端整体, cols 11-15 = APP, cols 17-21 = Web
    # Platform column mapping: {platform: {db_obj_col, db_sub_col, qw_obj_col, qw_sub_col}}
    plat_cols = {
        '双端整体': {'db_obj': 6, 'db_sub': 7, 'qw_obj': 8, 'qw_sub': 9},
        'APP':      {'db_obj': 12, 'db_sub': 13, 'qw_obj': 14, 'qw_sub': 15},
        'Web':      {'db_obj': 18, 'db_sub': 19, 'qw_obj': 20, 'qw_sub': 21},
    }
    # Word count has only 2 cols per platform (no obj/sub split)
    plat_wc_cols = {
        '双端整体': {'db': 6, 'qw': 8},
        'APP':      {'db': 12, 'qw': 14},
        'Web':      {'db': 18, 'qw': 20},
    }

    metrics_order = ['GSB', 'p-value', '可用率', '满分率', '劝退率', '均分', '回复字数']

    def read_section_block(start_row):
        """Read a 7-row metrics block starting at start_row (0-indexed).
        Returns dict: {platform: {metric: {key: val}}}
        """
        result = {}
        for plat in ['双端整体', 'APP', 'Web']:
            pc = plat_cols[plat]
            wc = plat_wc_cols[plat]
            pdata = {}
            for mi, metric in enumerate(metrics_order):
                ri = start_row + mi
                if ri >= nrows:
                    continue
                r = all_rows[ri]
                def cv(c):
                    return r[c] if c < len(r) else None
                if metric == '回复字数':
                    pdata[metric] = {
                        '豆包': fmt_wc(cv(wc['db'])),
                        'Qwen': fmt_wc(cv(wc['qw']))
                    }
                elif metric == 'p-value':
                    pdata[metric] = {
                        '豆包_客观': fmt_pval(cv(pc['db_obj'])),
                        '豆包_主客观': fmt_pval(cv(pc['db_sub'])),
                        'Qwen_客观': fmt_pval(cv(pc['qw_obj'])),
                        'Qwen_主客观': fmt_pval(cv(pc['qw_sub']))
                    }
                else:
                    pdata[metric] = {
                        '豆包_客观': fmt_pct(cv(pc['db_obj'])),
                        '豆包_主客观': fmt_pct(cv(pc['db_sub'])),
                        'Qwen_客观': fmt_pct(cv(pc['qw_obj'])),
                        'Qwen_主客观': fmt_pct(cv(pc['qw_sub']))
                    }
            result[plat] = pdata
        return result

    stats = {'双端整体': {}, 'APP': {}, 'Web': {}}

    # --- Session dimension ---
    # Row 5 (0-indexed): section title "整体（session维度）" in col 5
    # Row 6: headers
    # Row 7: metrics start (GSB)
    # But from our data dump: Row 8 (1-indexed) = Row 7 (0-indexed) is GSB for session
    session_metrics_start = 7  # 0-indexed row for GSB
    session_block = read_section_block(session_metrics_start)
    for plat in ['双端整体', 'APP', 'Web']:
        stats[plat]['session'] = {'整体': session_block[plat]}

    # --- Prompt dimension: 整体（prompt维度）---
    # Row 17 (1-indexed) = Row 16 (0-indexed): section title
    # Row 18: headers, Row 19: sub-headers
    # Row 20 (1-indexed) = Row 19 (0-indexed): GSB start
    prompt_overall_start = 19  # 0-indexed
    prompt_overall_block = read_section_block(prompt_overall_start)

    # --- Category sections in prompt dimension ---
    # Each section: title row, header row, sub-header row, then 7 metric rows
    # Sections start at rows (1-indexed): 29, 41, 53, 65, 77, 89, 101, 113, 125, 137, 149
    # = (0-indexed): 28, 40, 52, 64, 76, 88, 100, 112, 124, 136, 148
    # Metrics start 3 rows after title: title + header + sub-header

    # Map section titles to category names used in DATA
    section_title_to_cat = {
        '知识联网+非联网': '知识联网+非联网',
        'VLM通用': 'VLM通用',
        '写作': '创作',  # Excel says "写作" but old DATA uses "创作"
        '角色扮演/对话': '角色扮演/对话',
        '多模态需求': '多模态需求',
        'VLM教育': 'VLM教育',
        '医疗': '医疗',
        '翻译': '翻译',
        '电商': '电商',
        '代码': '代码',
        '本地生活': '本地生活',
    }

    # Find all section blocks by scanning for title rows in col 5
    cat_blocks = {}
    for ri in range(nrows):
        r = all_rows[ri]
        v5 = r[5] if len(r) > 5 else None
        if v5 and isinstance(v5, str) and '（prompt维度）' in v5 and v5 != '整体（prompt维度）':
            # Extract category name: "知识联网+非联网（prompt维度）" -> "知识联网+非联网"
            cat_raw = v5.replace('（prompt维度）', '')
            cat_name = section_title_to_cat.get(cat_raw, cat_raw)
            metrics_start = ri + 3  # skip title, header, sub-header rows
            cat_blocks[cat_name] = read_section_block(metrics_start)

    # Build prompt stats for each platform
    for plat in ['双端整体', 'APP', 'Web']:
        prompt_data = {'整体': prompt_overall_block[plat]}
        for cat_name, block in cat_blocks.items():
            prompt_data[cat_name] = block[plat]
        stats[plat]['prompt'] = prompt_data

    data = {'stats': stats, 'volume': volume}
    print('Dashboard DATA built: %d platforms, prompt categories: %s' % (
        len(stats), list(cat_blocks.keys())))
    return data


# ============================================================
# STEP 3: Build HTML
# ============================================================
def build_html(cases, dashboard_data, base_html_path, case_js_path, output_path):
    with open(base_html_path, 'r', encoding='utf-8') as f:
        html = f.read()
    with open(case_js_path, 'r', encoding='utf-8') as f:
        case_logic_js = f.read()

    # Fix GSB display in dashboard (remove negation)
    html = html.replace(
        "var neg = -v;\n    return (neg>0?'+':'')+neg.toFixed(2)+'%';",
        "return (v>0?'+':'')+v.toFixed(2)+'%';"
    )
    html = html.replace(
        "return (-v)>=0?'green':'red';",
        "return v>=0?'green':'red';"
    )
    html = html.replace("var nv1=(v1!==null)?-v1:null;", "var nv1=v1;")
    html = html.replace("var nv2=(v2!==null)?-v2:null;", "var nv2=v2;")

    # Add annotation CSS
    ann_css = """<style>
.ann-block{padding:6px 14px 10px;font-size:12px;line-height:1.7;color:#4e5969;border-bottom:1px dashed #e5e6eb}
.ann-line{margin-bottom:2px}
.ann-lbl{color:#86909c;font-weight:500}
.ann-rich{margin-top:4px;padding:6px 10px;background:#fffbe6;border-radius:6px;border:1px solid #ffe58f}
.gsb-ann{padding:8px 20px 10px;font-size:12px;line-height:1.7;color:#4e5969;background:#fafbfc}
.gsb-ann .ann-lbl{color:#86909c;font-weight:500}
</style>"""
    html = html.replace('</head>', ann_css + '\n</head>')

    # ---- Replace var DATA (if dashboard_data provided) ----
    if dashboard_data is not None:
        data_start = html.find('var DATA = {')
        assert data_start >= 0, 'var DATA not found in base HTML'
        depth = 0
        i = data_start + len('var DATA = ')
        data_obj_end = -1
        for j in range(i, len(html)):
            if html[j] == '{':
                depth += 1
            elif html[j] == '}':
                depth -= 1
            if depth == 0:
                data_obj_end = j + 1
                break
        assert data_obj_end > 0, 'Could not find end of DATA object'
        if html[data_obj_end] == ';':
            data_obj_end += 1
        new_data_str = 'var DATA = ' + json.dumps(dashboard_data, ensure_ascii=False) + ';'
        html = html[:data_start] + new_data_str + html[data_obj_end:]
        print('Replaced var DATA (%d chars)' % len(new_data_str))
    else:
        print('Keeping template var DATA (no dashboard data provided)')

    # ---- Replace CASES + case explorer logic ----
    cases_start = html.find('var CASES = [')
    assert cases_start >= 0, 'CASES not found in base HTML'
    m = re.search(r'filterCases\(\);\s*</script>', html[cases_start:])
    assert m, 'filterCases end not found'
    replace_end = cases_start + m.start() + len('filterCases();')

    cases_json_str = json.dumps(cases, ensure_ascii=False)
    cases_json_str = cases_json_str.replace('</script', '<\\/script').replace('</Script', '<\\/Script')
    new_block = 'var CASES = ' + cases_json_str + ';\n' + case_logic_js

    html = html[:cases_start] + new_block + html[replace_end:]

    # Verify
    for fn in ['filterCases', 'renderCase', 'var CASES', 'var DATA']:
        assert fn in html, fn + ' missing in output!'

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    size = os.path.getsize(output_path)
    print('Built: %s (%d bytes, %.1f MB)' % (output_path, size, size/1024/1024))
    return output_path


# ============================================================
# MAIN
# ============================================================
if __name__ == '__main__':
    excel_path = sys.argv[1] if len(sys.argv) > 1 else 'new_data.xlsx'
    base_html = 'viz_final_v6.html'
    case_js = 'case_explorer_v8.js'
    output = 'viz_output.html'

    print('Loading %s...' % excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # Parse cases
    cases = parse_cases(excel_path, wb)
    gsb_win = sum(1 for c in cases if c['gsb_cls'] == 'gsb-win')
    gsb_lose = sum(1 for c in cases if c['gsb_cls'] == 'gsb-lose')
    gsb_tie = sum(1 for c in cases if c['gsb_cls'] == 'gsb-tie')
    print('GSB distribution: win=%d tie=%d lose=%d' % (gsb_win, gsb_tie, gsb_lose))

    # Parse dashboard data (if sheet exists)
    if '数据统计' in wb.sheetnames:
        dashboard_data = parse_dashboard_data(wb)
    else:
        dashboard_data = None
        print('No 数据统计 sheet found — keeping template dashboard DATA')

    # Build HTML
    build_html(cases, dashboard_data, base_html, case_js, output)
    print('\nDone! Output: %s' % output)
